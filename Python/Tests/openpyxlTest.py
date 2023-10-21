import openpyxl as xl
fileName = input("请输入文件名：\n")+".xlsx"
wb = xl.load_workbook(fileName)
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    prize = sheet.cell(row, 3)
    discount = sheet.cell(row, 4)
    real_prize = sheet.cell(row, 5)
    if "￥" or "$" in str(prize.value):
        re_prize = str(prize.value)[1:]
    else:
        re_prize = prize.value
    if "-" in str(discount.value):
        re_discount = 1 + discount.value
    else:
        re_discount = discount.value
    real_prize.value = float(re_prize) * float(re_discount)

wb.save('testXl2.xlsx')
print("successful")
