import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('testXl.xlsx')  # 加载 testXl.xlsx 并返回一个工作簿(workBook)对象
sheet = wb['Sheet1']  # 赋值表格1(Sheet1)给 sheet , 注意这是固定写法

cell01 = sheet['a1']  # 赋值之前定义的表格的a1位置给 cell01
# cell01 = sheet.cell(1, 1) 或者这样也可以

print(cell01.value)  # 打印cell01的值
print(sheet.max_row)  # 打印sheet的列数

wb.save('textXl02')  # 保存并命名为textXl02
